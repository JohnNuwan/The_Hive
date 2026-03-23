"""Audite le biais directionnel du champion scalp live.

Ce script reste strictement local. Il permet de :
    - reproduire le constat live a partir d'un export MT5 Excel ;
    - auditer hors ligne le checkpoint live actuel avec les metriques Arena.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
for import_root in (REPO_ROOT / "src" / "eva-lab", REPO_ROOT / "src" / "shared"):
    sys.path.insert(0, str(import_root))


def _to_float(value: Any, default: float = 0.0) -> float:
    """Convertit une valeur heterogene en flottant.

    Args:
        value (Any): Valeur brute.
        default (float): Valeur de repli.

    Returns:
        float: Valeur convertie.
    """
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _find_deal_rows(sheet) -> list[dict[str, Any]]:
    """Extrait les lignes de deals depuis l'export Excel MT5.

    Args:
        sheet: Feuille Excel active.

    Returns:
        list[dict[str, Any]]: Lignes normalisees de deals.
    """
    deal_rows: list[dict[str, Any]] = []
    date_pattern = re.compile(r"^\d{4}\.\d{2}\.\d{2} ")

    for row in sheet.iter_rows(values_only=True):
        if not row or len(row) < 14:
            continue
        time_value = row[0]
        order_type = str(row[3] or "").strip().lower()
        direction = str(row[4] or "").strip().lower()
        if not isinstance(time_value, str) or not date_pattern.match(time_value):
            continue
        if order_type not in {"buy", "sell"} or direction not in {"in", "out"}:
            continue

        deal_rows.append(
            {
                "time": time_value,
                "symbol": str(row[2] or "").strip(),
                "type": order_type,
                "direction": direction,
                "volume": _to_float(row[5]),
                "price": _to_float(row[6]),
                "order": row[7],
                "commission": _to_float(row[8]),
                "swap": _to_float(row[9]),
                "profit": _to_float(row[11]),
                "balance": _to_float(row[12]),
                "comment": str(row[13] or "").strip(),
            }
        )

    return deal_rows


def _normalize_label(value: Any) -> str:
    """Normalise un libelle Excel pour comparaison robuste.

    Args:
        value (Any): Libelle brut.

    Returns:
        str: Libelle normalise en ASCII minuscule.
    """
    raw_value = str(value or "").strip().lower()
    return unicodedata.normalize("NFKD", raw_value).encode("ascii", "ignore").decode("ascii")


def _extract_summary(sheet) -> dict[str, Any]:
    """Lit les metriques principales depuis la section ``Resultats``.

    Args:
        sheet: Feuille Excel active.

    Returns:
        dict[str, Any]: Resume exploitable du challenge.
    """
    summary_map: dict[str, Any] = {}
    label_pairs = {
        "profit total net:": "profit_total_net",
        "facteur de profit:": "profit_factor",
        "solde drawdown maximal:": "max_drawdown",
        "nb trades:": "total_trades",
        "positions courtes (gagnees %):": "short_positions",
        "positions longues (gagnees %):": "long_positions",
    }

    for row in sheet.iter_rows(values_only=True):
        values = list(row)
        for index, value in enumerate(values):
            label = _normalize_label(value)
            if label not in label_pairs:
                continue
            key = label_pairs[label]
            extracted = values[index + 3] if index + 3 < len(values) else None
            summary_map[key] = extracted

    return summary_map


def build_excel_audit(
    workbook_path: Path,
    comment_prefix: str,
    target_day: str | None,
) -> dict[str, Any]:
    """Construit le rapport d'audit live depuis l'export MT5.

    Args:
        workbook_path (Path): Export Excel du challenge.
        comment_prefix (str): Prefixe des ordres automatiques scalp.
        target_day (str | None): Jour a isoler, format ``YYYY-MM-DD``.

    Returns:
        dict[str, Any]: Rapport de biais directionnel live.
    """
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    deals = _find_deal_rows(sheet)
    summary = _extract_summary(sheet)

    auto_entries = [
        row
        for row in deals
        if row["direction"] == "in" and row["comment"].startswith(comment_prefix)
    ]
    auto_entries_by_side = Counter(row["type"].upper() for row in auto_entries)

    target_rows = []
    target_pnl_by_symbol: dict[str, float] = defaultdict(float)
    if target_day:
        normalized_day = datetime.fromisoformat(target_day).strftime("%Y.%m.%d")
        target_rows = [
            row for row in deals if row["direction"] == "out" and str(row["time"]).startswith(normalized_day)
        ]
        for row in target_rows:
            target_pnl_by_symbol[row["symbol"]] += float(row["profit"])

    return {
        "workbook": str(workbook_path),
        "comment_prefix": comment_prefix,
        "summary": {
            "profit_total_net": _to_float(summary.get("profit_total_net")),
            "profit_factor": _to_float(summary.get("profit_factor")),
            "max_drawdown": str(summary.get("max_drawdown") or ""),
            "total_trades": int(_to_float(summary.get("total_trades"))),
            "short_positions": str(summary.get("short_positions") or ""),
            "long_positions": str(summary.get("long_positions") or ""),
        },
        "auto_entries": {
            "total": len(auto_entries),
            "buy": int(auto_entries_by_side.get("BUY", 0)),
            "sell": int(auto_entries_by_side.get("SELL", 0)),
        },
        "target_day": {
            "date": target_day,
            "closed_deals": len(target_rows),
            "realized_pnl": round(sum(row["profit"] for row in target_rows), 2),
            "pnl_by_symbol": {
                symbol: round(value, 2)
                for symbol, value in sorted(target_pnl_by_symbol.items())
            },
            "auto_entries": {
                "buy": sum(
                    1
                    for row in auto_entries
                    if target_day
                    and str(row["time"]).startswith(datetime.fromisoformat(target_day).strftime("%Y.%m.%d"))
                    and row["type"] == "buy"
                ),
                "sell": sum(
                    1
                    for row in auto_entries
                    if target_day
                    and str(row["time"]).startswith(datetime.fromisoformat(target_day).strftime("%Y.%m.%d"))
                    and row["type"] == "sell"
                ),
            },
        },
    }


def build_model_audit(horizon: str) -> dict[str, Any]:
    """Construit un audit hors ligne du checkpoint live courant.

    Args:
        horizon (str): Horizon cible.

    Returns:
        dict[str, Any]: Metriques directionnelles Arena du checkpoint live.
    """
    from eva_lab.arena import Arena  # Import paresseux pour laisser l'audit Excel autonome.
    from eva_lab.champion_promoter import ChampionPromoter

    promoter = ChampionPromoter()
    checkpoint_path, live_meta = promoter.resolve_live_checkpoint(horizon)
    if checkpoint_path is None:
        return {
            "status": "unavailable",
            "reason": "aucun_checkpoint_live",
            "horizon": horizon,
        }

    live_universe = promoter.build_live_universe(horizon)
    symbols = list(live_universe.get("symbols", []) or [])
    if not symbols:
        symbols = ["XAUUSD", "EURUSD", "GBPUSD", "USDJPY", "BTCUSD"]

    arena = Arena()
    metrics = arena._evaluate_model(checkpoint_path, symbols, horizon)
    return {
        "status": "ok",
        "horizon": horizon,
        "checkpoint": str(checkpoint_path),
        "selection": live_meta.get("selection"),
        "symbols": symbols,
        "metrics": metrics,
    }


def parse_args() -> argparse.Namespace:
    """Construit l'interface en ligne de commande.

    Returns:
        argparse.Namespace: Arguments valides.
    """
    parser = argparse.ArgumentParser(description="Audit local du biais directionnel du scalp live.")
    parser.add_argument(
        "--excel-path",
        type=Path,
        default=Path(r"C:\Users\nandi\Desktop\challenge.xlsx"),
        help="Chemin de l'export Excel MT5.",
    )
    parser.add_argument(
        "--comment-prefix",
        default="MZ-SCP-",
        help="Prefixe des commentaires a auditer pour les entrees scalp automatiques.",
    )
    parser.add_argument(
        "--date",
        default="2026-03-16",
        help="Jour a isoler au format YYYY-MM-DD.",
    )
    parser.add_argument(
        "--with-model-audit",
        action="store_true",
        help="Ajoute un audit Arena du checkpoint live courant.",
    )
    parser.add_argument(
        "--horizon",
        default="scalp",
        help="Horizon MuZero a auditer pour le checkpoint live.",
    )
    return parser.parse_args()


def main() -> int:
    """Point d'entree du script."""
    args = parse_args()
    report = {
        "generated_at": datetime.now().isoformat(),
        "excel_audit": build_excel_audit(args.excel_path, args.comment_prefix, args.date),
    }
    if args.with_model_audit:
        report["model_audit"] = build_model_audit(args.horizon)

    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
